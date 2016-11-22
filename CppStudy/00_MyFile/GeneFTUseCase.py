#!/usr/bin/env python
# coding=utf-8

import os
import sys
import re
import subprocess
import time
import openpyxl as xls

teamName="CRTeam2"

AllUseCase=[]
def delPause(filepath) :
  pfileIn=open(filepath+"run_eFT.bat",'r')
  pfileOut=open(filepath+"run_eFTtmp.bat",'w')
  
  for line in pfileIn.readlines():
    if "pause" != line.strip('\n'):
      pfileOut.write(line)
    else :
      print(line + " is not write new file")
  pfileIn.close()
  pfileOut.close()
  
def GetTestCase2File(InFIleName):
  global AllUseCase
  print("GetTestCase2File:input file is \"%s\"" % (InFIleName))
  pfileIn=open(InFIleName,'r')
  count=0
  for line in pfileIn.readlines():
    if re.search("testcase name=",line):
#      pfileOut.write("\"%s\",\"%s\",\"%s\"\n" % (line.split('"')[7],line.split('"')[3],line.split('"')[1]))
      tmp = {'FileName':line.split('"')[7],'Fiture':line.split('"')[3],'UsecaseName':line.split('"')[1]}
      count += 1
      AllUseCase.append(tmp)
      '''print("AllUseCase len is %d,%s,%s,%s" % (len(AllUseCase),
                                               AllUseCase[len(AllUseCase)-1]['FileName'],
                                               AllUseCase[len(AllUseCase)-1]['Fiture'],
                                               AllUseCase[len(AllUseCase)-1]['UsecaseName']))
      '''
  pfileIn.close()
  print("GetTestCase2File is end, %d\n" % count)
  
def main():
  global AllUseCase
  rootpath=sys.path[0]
  eFTPath=rootpath + '\\code\\Code_LTE\\cmac\\eFT\\'
  delPause(eFTPath);
  run_eFTForTeam = eFTPath+"run_eFTtmp.bat"+' -t:' + teamName
  print("command is %s" % run_eFTForTeam)
  subprocess.check_call(run_eFTForTeam)
  #print('del /Q /F '+eFTPath+"run_eFTtmp.bat")
  os.remove(eFTPath+"run_eFTtmp.bat")

  OutFileName=rootpath+"\\"+teamName+"eFTUseCase"+time.strftime("%Y%m%d_%X",time.localtime(time.time())).replace(':','')+".xlsx"
  GetTestCase2File(rootpath+'\\code\\Code_LTE\\cmac\\eFT\\eFT\\Debug\\testresult_demo.xml')
  GetTestCase2File(rootpath+'\\code\\Code_LTE\\cmac\\eFT\\eFT\\Debug\\testresult_feature.xml')
  GetTestCase2File(rootpath+'\\code\\Code_LTE\\cmac\\eFT\\eFT\\Debug\\testresult_module.xml')
  GetTestCase2File(rootpath+'\\code\\Code_LTE\\cmac\\eFT\\eFT\\Debug\\testresult_pressure.xml')
  GetTestCase2File(rootpath+'\\code\\Code_LTE\\cmac\\eFT\\eFT\\Debug\\testresult_radio.xml')
  GetTestCase2File(rootpath+'\\code\\Code_LTE\\cmac\\eFT\\eFT\\Debug\\testresult_unit.xml')
  
  wb=xls.workbook.Workbook()
  ws=wb.get_active_sheet()
  ws.title='AlleFTUseCase'
  
  UseCaseStatic={}
  RawLine = 0
  ws.cell(row=RawLine, column=0).value = 'FileName';
  ws.cell(row=RawLine, column=1).value = 'Fiture';
  ws.cell(row=RawLine, column=2).value = 'UsecaseName';
  for aUseCase in AllUseCase:
    RawLine += 1
    ws.cell(row=RawLine, column=0).value = aUseCase['FileName'];
    ws.cell(row=RawLine, column=1).value = aUseCase['Fiture'];
    ws.cell(row=RawLine, column=2).value = aUseCase['UsecaseName'];
    if UseCaseStatic.has_key(aUseCase['Fiture']) :
      UseCaseStatic[aUseCase['Fiture']] += 1
    else:
      UseCaseStatic[aUseCase['Fiture']] = 1
  
  print("UseCaseStatic is \n", UseCaseStatic)
  
  ws2=wb.create_sheet(index=1)
  ws2.title='Static'
  RawLine = 0
  ws2.cell(row=RawLine,column=0).value = 'Fiture'
  ws2.cell(row=RawLine,column=1).value = 'UsecaseNameNum'
  for key in UseCaseStatic.keys():
    RawLine += 1
    ws2.cell(row=RawLine,column=0).value = key
    ws2.cell(row=RawLine,column=1).value = UseCaseStatic[key]
    
  #chart=LineChart()
    
  #ws2.add_chart(chart)
  wb.save(OutFileName)
  

if __name__ == '__main__':
  main()



